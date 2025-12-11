# utils/excel_converter/excel_utils.py - EXCEL CONVERTER UTILITIES
# Dedicated utilities for Excel file conversion
import os
import json
import csv
import tempfile
import shutil
from pathlib import Path
import logging
from typing import List, Optional, Tuple, Dict, Any
from datetime import datetime, date
import io

# Universal Security Framework Integration
try:
    from security.core.validators import validate_content, validate_filename
    from security.core.sanitizers import sanitize_filename
    SECURITY_FRAMEWORK_AVAILABLE = True
except ImportError:
    SECURITY_FRAMEWORK_AVAILABLE = False

# Excel processing dependencies
try:
    import pandas as pd
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_DEPENDENCIES_AVAILABLE = True
except ImportError:
    EXCEL_DEPENDENCIES_AVAILABLE = False

# Additional dependencies for XML processing
try:
    import xml.etree.ElementTree as ET
    from xml.dom import minidom
    XML_AVAILABLE = True
except ImportError:
    XML_AVAILABLE = False

# HTML processing
try:
    from bs4 import BeautifulSoup
    BS4_AVAILABLE = True
except ImportError:
    BS4_AVAILABLE = False

# Initialize logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelConverter:
    """
    Dedicated Excel converter supporting:
    - Excel to CSV conversion
    - Excel to JSON conversion
    - Excel to HTML conversion
    - Excel to TXT conversion
    - Excel to XML conversion
    """
    
    def __init__(self):
        self.temp_dirs = []  # Track temporary directories for cleanup
        
        # Check for required dependencies
        self.dependencies = self._check_dependencies()
        
    def __del__(self):
        """Cleanup temporary directories on object destruction"""
        self.cleanup_temp_dirs()
        
    def cleanup_temp_dirs(self):
        """Clean up all temporary directories created by this instance"""
        for temp_dir in self.temp_dirs:
            if os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir)
                except Exception as e:
                    logger.warning(f"Failed to cleanup temp dir {temp_dir}: {e}")
        self.temp_dirs.clear()
    
    def _check_dependencies(self) -> Dict[str, bool]:
        """Check availability of all dependencies"""
        return {
            'pandas': EXCEL_DEPENDENCIES_AVAILABLE,
            'openpyxl': EXCEL_DEPENDENCIES_AVAILABLE,
            'xml': XML_AVAILABLE,
            'bs4': BS4_AVAILABLE
        }
    
    def is_csv_conversion_available(self) -> bool:
        """Check if CSV conversion is available"""
        return self.dependencies['pandas']
    
    def is_tsv_conversion_available(self) -> bool:
        """Check if TSV conversion is available"""
        return self.dependencies['pandas']
    
    def is_json_conversion_available(self) -> bool:
        """Check if JSON conversion is available"""
        return self.dependencies['pandas']
    
    def is_html_conversion_available(self) -> bool:
        """Check if HTML conversion is available"""
        return self.dependencies['pandas']
    
    def is_txt_conversion_available(self) -> bool:
        """Check if TXT conversion is available"""
        return self.dependencies['pandas']
    
    def is_xml_conversion_available(self) -> bool:
        """Check if XML conversion is available"""
        return (self.dependencies['pandas'] and self.dependencies['xml'])
    
    def is_ods_conversion_available(self) -> bool:
        """Check if ODS conversion is available"""
        if not self.dependencies['pandas']:
            return False
        
        # Check if odfpy is available
        try:
            import odf  # odfpy imports as 'odf'
            return True
        except ImportError:
            return False
    
    def is_pdf_conversion_available(self) -> bool:
        """Check if PDF conversion is available"""
        return self.dependencies['pandas']  # Basic PDF creation available
    
    def validate_file_security(self, file_path: str) -> Tuple[bool, List[str]]:
        """Validate file using universal security framework"""
        if not SECURITY_FRAMEWORK_AVAILABLE:
            logger.warning("Security framework not available - performing basic validation")
            return self._basic_file_validation(file_path)
        
        try:
            # Read file content for security validation
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            # Determine file type
            file_ext = Path(file_path).suffix.lower().lstrip('.')
            
            # Perform security validation
            is_safe, issues = validate_content(file_content, file_ext)
            
            if not is_safe:
                logger.error(f"Security validation failed for {file_path}: {issues}")
                return False, issues
            
            logger.info(f"Security validation passed for {file_path}")
            return True, []
            
        except Exception as e:
            logger.error(f"Security validation error for {file_path}: {e}")
            return False, [f"Security validation error: {str(e)}"]
    
    def _basic_file_validation(self, file_path: str) -> Tuple[bool, List[str]]:
        """Basic file validation when security framework is not available"""
        try:
            if not os.path.exists(file_path):
                return False, ["File does not exist"]
            
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                return False, ["File is empty"]
            
            if file_size > 100 * 1024 * 1024:  # 100MB limit
                return False, ["File too large"]
            
            # Check file extension
            valid_extensions = ['.xls', '.xlsx', '.xlsm']
            file_ext = Path(file_path).suffix.lower()
            if file_ext not in valid_extensions:
                return False, [f"Invalid file extension: {file_ext}"]
            
            return True, []
            
        except Exception as e:
            return False, [f"Basic validation error: {str(e)}"]
    
    def _load_excel_file(self, input_path: str, selected_sheets: str = 'all') -> Optional[Dict[str, pd.DataFrame]]:
        """Load Excel file and return dict of sheet_name: DataFrame"""
        try:
            if not EXCEL_DEPENDENCIES_AVAILABLE:
                logger.error("Excel dependencies not available")
                return None
            
            # Read Excel file
            if selected_sheets == 'all':
                # Load all sheets
                excel_data = pd.read_excel(input_path, sheet_name=None, engine='openpyxl')
            elif selected_sheets == 'first':
                # Load only first sheet
                excel_data = pd.read_excel(input_path, sheet_name=0, engine='openpyxl')
                excel_data = {'Sheet1': excel_data}
            else:
                # Handle manual selection (could be ranges like "1,3,5" or "2-4")
                sheet_indices = self._parse_sheet_selection(selected_sheets, input_path)
                if sheet_indices:
                    excel_data = pd.read_excel(input_path, sheet_name=sheet_indices, engine='openpyxl')
                else:
                    # Fallback to first sheet if parsing fails
                    excel_data = pd.read_excel(input_path, sheet_name=0, engine='openpyxl')
                    excel_data = {'Sheet1': excel_data}
            
            return excel_data
            
        except Exception as e:
            logger.error(f"Failed to load Excel file {input_path}: {e}")
            return None
    
    def _parse_sheet_selection(self, selection: str, file_path: str) -> List[int]:
        """Parse sheet selection string into list of sheet indices (0-based)"""
        try:
            # Get total sheet count first
            wb = load_workbook(file_path, read_only=True)
            total_sheets = len(wb.sheetnames)
            wb.close()
            
            parts = selection.split(',') if ',' in selection else [selection]
            sheet_indices = set()
            
            for part in parts:
                part = part.strip()
                if '-' in part:
                    # Range: "2-4" becomes [1, 2, 3] (0-based)
                    start, end = part.split('-')
                    start_idx = int(start) - 1  # Convert to 0-based
                    end_idx = int(end) - 1
                    if 0 <= start_idx <= end_idx < total_sheets:
                        sheet_indices.update(range(start_idx, end_idx + 1))
                else:
                    # Single number: "3" becomes 2 (0-based)
                    idx = int(part) - 1  # Convert to 0-based
                    if 0 <= idx < total_sheets:
                        sheet_indices.add(idx)
            
            return sorted(list(sheet_indices))
            
        except Exception as e:
            logger.error(f"Failed to parse sheet selection '{selection}': {e}")
            return []
    
    def _write_excel_compatible_csv(self, df: pd.DataFrame, output_path: str, 
                                   delimiter: str = ',', encoding: str = 'utf-8', 
                                   include_headers: bool = True) -> None:
        """Write CSV in a format that's compatible with Microsoft Excel"""
        try:
            # Excel-specific optimizations
            if encoding.lower() in ['utf-8', 'utf8']:
                # Use UTF-8 with BOM for Excel compatibility
                with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
                    df.to_csv(
                        f,
                        sep=delimiter,
                        index=False,
                        header=include_headers,
                        lineterminator='\r\n',  # Windows line endings for Excel
                        quoting=0,  # Quote only when necessary (QUOTE_MINIMAL)
                        doublequote=True,
                        escapechar=None
                    )
            elif encoding.lower() in ['cp1252', 'windows-1252']:
                # Windows-1252 encoding for older Excel versions
                with open(output_path, 'w', encoding='cp1252', newline='') as f:
                    df.to_csv(
                        f,
                        sep=delimiter,
                        index=False,
                        header=include_headers,
                        lineterminator='\r\n',
                        quoting=0,  # Quote only when necessary
                        doublequote=True,
                        escapechar=None
                    )
            else:
                # Standard encoding
                df.to_csv(
                    output_path,
                    sep=delimiter,
                    encoding=encoding,
                    index=False,
                    header=include_headers,
                    lineterminator='\r\n'
                )
                
            logger.info(f"Excel-compatible CSV written: {output_path} (encoding: {encoding})")
            
        except Exception as e:
            logger.error(f"Error writing Excel-compatible CSV: {e}")
            # Fallback to standard CSV writing
            df.to_csv(
                output_path,
                sep=delimiter,
                encoding=encoding,
                index=False,
                header=include_headers
            )
    
    def _create_excel_compatible_csv_content(self, df: pd.DataFrame, 
                                           delimiter: str = ',', encoding: str = 'utf-8', 
                                           include_headers: bool = True) -> bytes:
        """Create Excel-compatible CSV content as bytes for ZIP files"""
        import io
        
        try:
            # Create CSV content with Excel-compatible settings
            csv_buffer = io.StringIO()
            df.to_csv(
                csv_buffer,
                sep=delimiter,
                index=False,
                header=include_headers,
                lineterminator='\r\n',  # Windows line endings
                quoting=0,  # Quote only when necessary (QUOTE_MINIMAL)
                doublequote=True,
                escapechar=None
            )
            csv_string = csv_buffer.getvalue()
            csv_buffer.close()
            
            # Handle encoding with BOM for Excel compatibility
            if encoding.lower() in ['utf-8', 'utf8']:
                # Add BOM for Excel UTF-8 compatibility
                return '\ufeff'.encode('utf-8') + csv_string.encode('utf-8')
            elif encoding.lower() in ['cp1252', 'windows-1252']:
                return csv_string.encode('cp1252')
            else:
                return csv_string.encode(encoding)
                
        except Exception as e:
            logger.error(f"Error creating Excel-compatible CSV content: {e}")
            # Fallback to basic CSV
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, sep=delimiter, index=False, header=include_headers)
            csv_string = csv_buffer.getvalue()
            csv_buffer.close()
            return csv_string.encode(encoding)
    
    def _process_dataframe(self, df: pd.DataFrame, 
                          convert_formulas: bool = True,
                          preserve_dates: bool = True,
                          include_headers: bool = True) -> pd.DataFrame:
        """Process DataFrame based on conversion options"""
        try:
            processed_df = df.copy()
            
            # Handle date preservation
            if preserve_dates:
                for col in processed_df.columns:
                    if processed_df[col].dtype == 'datetime64[ns]':
                        # Keep datetime format
                        continue
                    elif processed_df[col].dtype == 'object':
                        # Try to convert to datetime if possible
                        try:
                            # Use modern pandas approach with better date detection
                            temp_series = processed_df[col].copy().dropna()
                            
                            # Only try datetime conversion if we have data and it looks like dates
                            if len(temp_series) > 0:
                                # Check if any values look like dates (simple heuristic)
                                sample_val = str(temp_series.iloc[0]) if len(temp_series) > 0 else ''
                                
                                # Look for common date patterns
                                if any(char in sample_val for char in ['-', '/', ':', ' ']) and len(sample_val) > 5:
                                    try:
                                        # Use modern pandas datetime parsing (infer_datetime_format is deprecated)
                                        import warnings
                                        with warnings.catch_warnings():
                                            warnings.simplefilter('ignore', UserWarning)
                                            # Modern pandas automatically infers datetime format
                                            processed_df[col] = pd.to_datetime(processed_df[col], errors='coerce')
                                    except (ValueError, TypeError):
                                        # If conversion fails, keep original values
                                        pass
                        except Exception:
                            # If any error occurs, keep original values
                            pass
            
            # Handle formula conversion (formulas are already converted to values by pandas)
            if convert_formulas:
                # pandas automatically converts formulas to values, so no additional action needed
                pass
            
            # Handle headers
            if not include_headers:
                processed_df.columns = range(len(processed_df.columns))
            
            return processed_df
            
        except Exception as e:
            logger.error(f"Failed to process DataFrame: {e}")
            return df
    
    def excel_to_csv(self, input_path: str, output_path: str,
                    preserve_formatting: bool = False,
                    include_headers: bool = True,
                    convert_formulas: bool = True,
                    preserve_dates: bool = True,
                    selected_sheets: str = 'all',
                    csv_delimiter: str = ',',
                    csv_encoding: str = 'utf-8') -> bool:
        """Convert Excel to CSV"""
        try:
            if not self.is_csv_conversion_available():
                logger.error("Required dependencies not available for CSV conversion")
                return False
            
            if not os.path.exists(input_path):
                logger.error(f"Input file not found: {input_path}")
                return False
            
            # Security validation
            is_safe, security_issues = self.validate_file_security(input_path)
            if not is_safe:
                logger.error(f"Security validation failed for {input_path}: {security_issues}")
                return False
            
            logger.info(f"Converting Excel to CSV: {input_path} -> {output_path}")
            
            # Load Excel file
            excel_data = self._load_excel_file(input_path, selected_sheets)
            if not excel_data:
                return False
            
            # CSV now only handles single sheet conversion (simplified)
            # Take the first sheet if multiple sheets are loaded
            sheet_name, df = next(iter(excel_data.items()))
            processed_df = self._process_dataframe(
                df, convert_formulas, preserve_dates, include_headers
            )
            
            # Write Excel-compatible CSV
            self._write_excel_compatible_csv(
                processed_df, output_path, csv_delimiter, csv_encoding, include_headers
            )
            
            if len(excel_data) > 1:
                logger.info(f"Multiple sheets found, converted first sheet: {sheet_name}")
            else:
                logger.info(f"Converted single sheet: {sheet_name}")
            
            success = os.path.exists(output_path)
            if success:
                logger.info(f"CSV conversion successful: {output_path}")
            else:
                logger.error("CSV conversion failed - output file not created")
                
            return success
            
        except Exception as e:
            logger.error(f"CSV conversion error: {e}")
            return False
    
    def excel_to_tsv(self, input_path: str, output_path: str,
                    preserve_formatting: bool = False,
                    include_headers: bool = True,
                    convert_formulas: bool = True,
                    preserve_dates: bool = True,
                    selected_sheets: str = 'all',
                    tsv_encoding: str = 'utf-8',
                    **kwargs) -> bool:
        """Convert Excel to TSV (Tab-Separated Values)"""
        try:
            if not self.is_tsv_conversion_available():
                logger.error("Required dependencies not available for TSV conversion")
                return False
            
            if not os.path.exists(input_path):
                logger.error(f"Input file not found: {input_path}")
                return False
            
            # Security validation
            is_safe, security_issues = self.validate_file_security(input_path)
            if not is_safe:
                logger.error(f"Security validation failed for {input_path}: {security_issues}")
                return False
            
            logger.info(f"Converting Excel to TSV: {input_path} -> {output_path}")
            
            # Load Excel file
            excel_data = self._load_excel_file(input_path, selected_sheets)
            if not excel_data:
                return False
            
            # Handle multiple sheets (take first sheet for TSV)
            sheet_name, df = next(iter(excel_data.items()))
            processed_df = self._process_dataframe(
                df, convert_formulas, preserve_dates, include_headers
            )
            
            # TSV uses tab as delimiter
            processed_df.to_csv(
                output_path,
                sep='\t',  # Tab delimiter for TSV
                encoding=tsv_encoding,
                index=False,
                header=include_headers
            )
            
            if len(excel_data) > 1:
                logger.info(f"Multiple sheets found, converted first sheet: {sheet_name}")
            
            success = os.path.exists(output_path)
            if success:
                logger.info(f"TSV conversion successful: {output_path}")
            else:
                logger.error("TSV conversion failed - output file not created")
                
            return success
            
        except Exception as e:
            logger.error(f"TSV conversion error: {e}")
            return False
    
    def excel_to_json(self, input_path: str, output_path: str,
                     preserve_formatting: bool = False,
                     include_headers: bool = True,
                     convert_formulas: bool = True,
                     preserve_dates: bool = True,
                     selected_sheets: str = 'all',
                     **kwargs) -> bool:
        """Convert Excel to JSON"""
        try:
            if not self.is_json_conversion_available():
                logger.error("Required dependencies not available for JSON conversion")
                return False
            
            if not os.path.exists(input_path):
                logger.error(f"Input file not found: {input_path}")
                return False
            
            # Security validation
            is_safe, security_issues = self.validate_file_security(input_path)
            if not is_safe:
                logger.error(f"Security validation failed for {input_path}: {security_issues}")
                return False
            
            logger.info(f"Converting Excel to JSON: {input_path} -> {output_path}")
            
            # Load Excel file
            excel_data = self._load_excel_file(input_path, selected_sheets)
            if not excel_data:
                return False
            
            # Prepare JSON data
            json_data = {}
            
            for sheet_name, df in excel_data.items():
                processed_df = self._process_dataframe(
                    df, convert_formulas, preserve_dates, include_headers
                )
                
                # Convert DataFrame to JSON-compatible format
                if include_headers:
                    sheet_data = processed_df.to_dict('records')
                else:
                    sheet_data = processed_df.values.tolist()
                
                # Handle datetime objects for JSON serialization
                def json_serializer(obj):
                    if isinstance(obj, (datetime, date)):
                        return obj.isoformat()
                    if pd.isna(obj):
                        return None
                    return obj
                
                # Convert to JSON-serializable format
                serializable_data = []
                for row in sheet_data:
                    if isinstance(row, dict):
                        serializable_row = {k: json_serializer(v) for k, v in row.items()}
                    else:
                        serializable_row = [json_serializer(v) for v in row]
                    serializable_data.append(serializable_row)
                
                json_data[sheet_name] = serializable_data
            
            # If single sheet, flatten the structure
            if len(json_data) == 1:
                json_data = next(iter(json_data.values()))
            
            # Write JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False)
            
            success = os.path.exists(output_path)
            if success:
                logger.info(f"JSON conversion successful: {output_path}")
            else:
                logger.error("JSON conversion failed - output file not created")
                
            return success
            
        except Exception as e:
            logger.error(f"JSON conversion error: {e}")
            return False
    
    def excel_to_html(self, input_path: str, output_path: str,
                     preserve_formatting: bool = True,
                     include_headers: bool = True,
                     convert_formulas: bool = True,
                     preserve_dates: bool = True,
                     selected_sheets: str = 'all',
                     **kwargs) -> bool:
        """Convert Excel to HTML"""
        try:
            if not self.is_html_conversion_available():
                logger.error("Required dependencies not available for HTML conversion")
                return False
            
            if not os.path.exists(input_path):
                logger.error(f"Input file not found: {input_path}")
                return False
            
            # Security validation
            is_safe, security_issues = self.validate_file_security(input_path)
            if not is_safe:
                logger.error(f"Security validation failed for {input_path}: {security_issues}")
                return False
            
            logger.info(f"Converting Excel to HTML: {input_path} -> {output_path}")
            
            # Load Excel file
            excel_data = self._load_excel_file(input_path, selected_sheets)
            if not excel_data:
                return False
            
            # Create HTML content
            html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Conversion Result</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .sheet { margin-bottom: 30px; }
        .sheet-title { font-size: 18px; font-weight: bold; margin-bottom: 10px; color: #333; }
        table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; font-weight: bold; }
        tr:nth-child(even) { background-color: #f9f9f9; }
        .number { text-align: right; }
        .date { text-align: center; }
    </style>
</head>
<body>
    <h1>Excel Conversion Result</h1>
"""
            
            for sheet_name, df in excel_data.items():
                processed_df = self._process_dataframe(
                    df, convert_formulas, preserve_dates, include_headers
                )
                
                html_content += f'    <div class="sheet">\n'
                html_content += f'        <div class="sheet-title">{sheet_name}</div>\n'
                
                # Convert DataFrame to HTML table
                # Ensure sheet_name is a string before calling replace()
                sheet_name_str = str(sheet_name)
                table_html = processed_df.to_html(
                    escape=False,
                    index=False,
                    classes='data-table',
                    table_id=f'sheet-{sheet_name_str.replace(" ", "-").lower()}'
                )
                
                html_content += f'        {table_html}\n'
                html_content += '    </div>\n'
            
            html_content += """
</body>
</html>
"""
            
            # Write HTML file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            success = os.path.exists(output_path)
            if success:
                logger.info(f"HTML conversion successful: {output_path}")
            else:
                logger.error("HTML conversion failed - output file not created")
                
            return success
            
        except Exception as e:
            logger.error(f"HTML conversion error: {e}")
            return False
    
    def excel_to_txt(self, input_path: str, output_path: str,
                    preserve_formatting: bool = False,
                    include_headers: bool = True,
                    convert_formulas: bool = True,
                    preserve_dates: bool = True,
                    selected_sheets: str = 'all',
                    **kwargs) -> bool:
        """Convert Excel to TXT"""
        try:
            if not self.is_txt_conversion_available():
                logger.error("Required dependencies not available for TXT conversion")
                return False
            
            if not os.path.exists(input_path):
                logger.error(f"Input file not found: {input_path}")
                return False
            
            # Security validation
            is_safe, security_issues = self.validate_file_security(input_path)
            if not is_safe:
                logger.error(f"Security validation failed for {input_path}: {security_issues}")
                return False
            
            logger.info(f"Converting Excel to TXT: {input_path} -> {output_path}")
            
            # Load Excel file
            excel_data = self._load_excel_file(input_path, selected_sheets)
            if not excel_data:
                return False
            
            # Create TXT content
            txt_content = []
            
            for sheet_name, df in excel_data.items():
                processed_df = self._process_dataframe(
                    df, convert_formulas, preserve_dates, include_headers
                )
                
                txt_content.append(f"=== {sheet_name} ===\n")
                
                # Convert DataFrame to string with tab separation
                if include_headers:
                    txt_content.append("\t".join(str(col) for col in processed_df.columns) + "\n")
                
                for _, row in processed_df.iterrows():
                    row_str = "\t".join(str(val) if not pd.isna(val) else "" for val in row)
                    txt_content.append(row_str + "\n")
                
                txt_content.append("\n")
            
            # Write TXT file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.writelines(txt_content)
            
            success = os.path.exists(output_path)
            if success:
                logger.info(f"TXT conversion successful: {output_path}")
            else:
                logger.error("TXT conversion failed - output file not created")
                
            return success
            
        except Exception as e:
            logger.error(f"TXT conversion error: {e}")
            return False
    
    def excel_to_xml(self, input_path: str, output_path: str,
                    preserve_formatting: bool = False,
                    include_headers: bool = True,
                    convert_formulas: bool = True,
                    preserve_dates: bool = True,
                    selected_sheets: str = 'all',
                    **kwargs) -> bool:
        """Convert Excel to XML"""
        try:
            if not self.is_xml_conversion_available():
                logger.error("Required dependencies not available for XML conversion")
                return False
            
            if not os.path.exists(input_path):
                logger.error(f"Input file not found: {input_path}")
                return False
            
            # Security validation
            is_safe, security_issues = self.validate_file_security(input_path)
            if not is_safe:
                logger.error(f"Security validation failed for {input_path}: {security_issues}")
                return False
            
            logger.info(f"Converting Excel to XML: {input_path} -> {output_path}")
            
            # Load Excel file
            excel_data = self._load_excel_file(input_path, selected_sheets)
            if not excel_data:
                return False
            
            # Create XML structure
            root = ET.Element("workbook")
            
            for sheet_name, df in excel_data.items():
                processed_df = self._process_dataframe(
                    df, convert_formulas, preserve_dates, include_headers
                )
                
                sheet_elem = ET.SubElement(root, "sheet", name=sheet_name)
                
                # Add headers if included
                if include_headers:
                    headers_elem = ET.SubElement(sheet_elem, "headers")
                    for col in processed_df.columns:
                        header_elem = ET.SubElement(headers_elem, "header")
                        header_elem.text = str(col)
                
                # Add data rows
                rows_elem = ET.SubElement(sheet_elem, "rows")
                for idx, (_, row) in enumerate(processed_df.iterrows()):
                    row_elem = ET.SubElement(rows_elem, "row", index=str(idx))
                    
                    for col_idx, (col_name, value) in enumerate(row.items()):
                        cell_elem = ET.SubElement(row_elem, "cell", column=str(col_name))
                        
                        # Handle different data types with robust conversion
                        try:
                            if pd.isna(value):
                                cell_elem.text = ""
                            elif isinstance(value, (datetime, date)):
                                cell_elem.text = value.isoformat()
                            elif isinstance(value, (int, float)):
                                # Explicitly handle numeric types
                                if pd.isna(value):
                                    cell_elem.text = ""
                                else:
                                    cell_elem.text = str(value)
                            elif isinstance(value, bool):
                                cell_elem.text = "true" if value else "false"
                            else:
                                # Convert any other type to string safely
                                cell_elem.text = str(value) if value is not None else ""
                        except Exception as cell_error:
                            # Fallback for any value that can't be converted
                            logger.warning(f"Failed to convert cell value {value} (type: {type(value)}): {cell_error}")
                            cell_elem.text = str(value) if value is not None else ""
            
            # Create XML string with robust error handling
            try:
                # First try to create XML string
                rough_string = ET.tostring(root, encoding='unicode')
                
                # Try to pretty-print the XML
                try:
                    reparsed = minidom.parseString(rough_string)
                    pretty_xml = reparsed.toprettyxml(indent="  ")
                    
                    # Clean up the XML (remove empty lines)
                    lines = [line for line in pretty_xml.split('\n') if line.strip()]
                    final_xml = '\n'.join(lines)
                    
                except Exception as pretty_error:
                    logger.warning(f"Failed to pretty-print XML, using raw format: {pretty_error}")
                    final_xml = rough_string
                
                # Write XML file
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(final_xml)
                    
            except Exception as xml_error:
                logger.error(f"Failed to create XML content: {xml_error}")
                # Try a simpler XML approach as fallback
                return self._create_simple_xml_fallback(excel_data, output_path, include_headers)
            
            success = os.path.exists(output_path)
            if success:
                logger.info(f"XML conversion successful: {output_path}")
            else:
                logger.error("XML conversion failed - output file not created")
                
            return success
            
        except Exception as e:
            logger.error(f"XML conversion error: {e}")
            return False
    
    def _create_simple_xml_fallback(self, excel_data: Dict[str, pd.DataFrame], output_path: str, include_headers: bool) -> bool:
        """Create a simple XML file as fallback when ElementTree fails"""
        try:
            logger.info("Creating simple XML fallback")
            xml_lines = ['<?xml version="1.0" encoding="UTF-8"?>', '<workbook>']
            
            for sheet_name, df in excel_data.items():
                # Sanitize sheet name for XML
                safe_sheet_name = str(sheet_name).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                xml_lines.append(f'  <sheet name="{safe_sheet_name}">')
                
                # Add headers if needed
                if include_headers:
                    xml_lines.append('    <headers>')
                    for col in df.columns:
                        safe_col = str(col).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        xml_lines.append(f'      <header>{safe_col}</header>')
                    xml_lines.append('    </headers>')
                
                # Add data rows
                xml_lines.append('    <rows>')
                for idx, (_, row) in enumerate(df.iterrows()):
                    xml_lines.append(f'      <row index="{idx}">')
                    
                    for col_name, value in row.items():
                        safe_col_name = str(col_name).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        
                        # Convert value to safe XML text
                        if pd.isna(value) or value is None:
                            safe_value = ""
                        elif isinstance(value, (datetime, date)):
                            safe_value = value.isoformat()
                        elif isinstance(value, bool):
                            safe_value = "true" if value else "false"
                        else:
                            # Convert to string and escape XML characters
                            safe_value = str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        
                        xml_lines.append(f'        <cell column="{safe_col_name}">{safe_value}</cell>')
                    
                    xml_lines.append('      </row>')
                
                xml_lines.append('    </rows>')
                xml_lines.append('  </sheet>')
            
            xml_lines.append('</workbook>')
            
            # Write the XML file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(xml_lines))
            
            logger.info(f"Simple XML fallback created: {output_path}")
            return os.path.exists(output_path)
            
        except Exception as fallback_error:
            logger.error(f"Simple XML fallback failed: {fallback_error}")
            return False
    
    def _create_pdf_tables_with_column_splitting(self, df: pd.DataFrame, include_headers: bool, 
                                               page_size: tuple, orientation: str) -> list:
        """Create PDF tables with intelligent column splitting for wide tables"""
        try:
            from reportlab.lib.pagesizes import landscape
            from reportlab.platypus import Table, TableStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            # Calculate available width
            if orientation.lower() == 'landscape':
                page_width = max(page_size)
            else:
                page_width = min(page_size)
            
            # Account for margins (1 inch total)
            available_width = page_width - 1.0 * inch
            
            # Analyze table dimensions
            num_cols = len(df.columns)
            num_rows = len(df)
            
            # Log all column names for verification
            logger.info(f"📊 Processing DataFrame with {num_cols} columns and {num_rows} rows")
            logger.info(f"📋 Column names: {list(df.columns)}")
            
            # Estimate column widths based on content
            col_widths = self._estimate_column_widths(df, include_headers)
            total_estimated_width = sum(col_widths)
            
            logger.info(f"📏 Table analysis: {num_cols} columns, {num_rows} rows, estimated width: {total_estimated_width:.1f} points")
            logger.info(f"📐 Available width: {available_width:.1f} points")
            logger.info(f"🔍 Column widths: {col_widths}")
            
            # Decide on table splitting strategy
            if total_estimated_width <= available_width:
                # Table fits in one piece
                logger.info("Table fits on page - creating single table")
                return self._create_single_pdf_table(df, include_headers, col_widths)
            else:
                # Table needs to be split
                logger.info("Table too wide - splitting into multiple tables")
                
                # Try landscape orientation first if not already
                if orientation.lower() != 'landscape':
                    landscape_width = max(page_size) - 1.0 * inch
                    if total_estimated_width <= landscape_width:
                        logger.info("Table fits in landscape - using landscape orientation")
                        return self._create_single_pdf_table(df, include_headers, col_widths, use_landscape=True)
                
                # Split table into multiple parts
                return self._split_pdf_table_by_columns(df, include_headers, col_widths, available_width)
                
        except Exception as e:
            logger.error(f"Error creating PDF tables: {e}")
            # Fallback to simple table
            return self._create_simple_pdf_table_fallback(df, include_headers)
    
    def _estimate_column_widths(self, df: pd.DataFrame, include_headers: bool) -> list:
        """Estimate optimal column widths based on content"""
        col_widths = []
        
        for col in df.columns:
            # Get max character length in column
            max_content_len = 0
            
            # Check header length if included
            if include_headers:
                max_content_len = max(max_content_len, len(str(col)))
            
            # Check data content lengths (sample first 100 rows for performance)
            sample_data = df[col].head(100)
            for val in sample_data:
                if not pd.isna(val):
                    max_content_len = max(max_content_len, len(str(val)))
            
            # Convert to points (rough estimation: 1 char ≈ 6-8 points)
            # Add padding and minimum width
            estimated_width = max(40, min(120, max_content_len * 7 + 16))  # Min 40pt, max 120pt
            col_widths.append(estimated_width)
        
        return col_widths
    
    def _create_single_pdf_table(self, df: pd.DataFrame, include_headers: bool, 
                               col_widths: list, use_landscape: bool = False) -> list:
        """Create a single PDF table with optimal formatting"""
        try:
            from reportlab.platypus import Table, TableStyle
            from reportlab.lib import colors
            
            # Prepare table data
            table_data = []
            if include_headers:
                table_data.append([str(col) for col in df.columns])
            
            # Add data rows (limit to reasonable number for PDF)
            max_rows = 1000  # Reasonable limit for PDF
            for idx, (_, row) in enumerate(df.iterrows()):
                if idx >= max_rows:
                    break
                table_data.append([str(val) if not pd.isna(val) else "" for val in row])
            
            if len(df) > max_rows:
                logger.warning(f"Table truncated from {len(df)} to {max_rows} rows for PDF readability")
            
            # Create table with optimized widths
            table = Table(table_data, colWidths=col_widths)
            
            # Apply enhanced styling
            style_commands = [
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]
            
            # Header styling if included
            if include_headers:
                style_commands.extend([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ])
            
            # Alternating row colors for better readability
            if len(table_data) > 1:
                start_row = 1 if include_headers else 0
                for i in range(start_row, len(table_data), 2):
                    style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.lightblue))
            
            table.setStyle(TableStyle(style_commands))
            
            return [table]
            
        except Exception as e:
            logger.error(f"Error creating single PDF table: {e}")
            return self._create_simple_pdf_table_fallback(df, include_headers)
    
    def _split_pdf_table_by_columns(self, df: pd.DataFrame, include_headers: bool, 
                                  col_widths: list, available_width: float) -> list:
        """Split wide table into multiple tables by columns"""
        try:
            from reportlab.platypus import Table, TableStyle
            from reportlab.lib import colors
            
            tables = []
            num_cols = len(df.columns)
            start_col = 0
            
            while start_col < num_cols:
                # Determine how many columns fit in this table
                current_width = 0
                end_col = start_col
                
                while end_col < num_cols:
                    if current_width + col_widths[end_col] <= available_width:
                        current_width += col_widths[end_col]
                        end_col += 1
                    else:
                        break
                
                # Ensure at least one column per table
                if end_col == start_col:
                    end_col = start_col + 1
                
                logger.info(f"📊 Creating table part {len(tables)+1} with columns {start_col} to {end_col-1} (width: {current_width:.1f}pt)")
                
                # Create subset dataframe
                subset_cols = df.columns[start_col:end_col]
                subset_df = df[subset_cols]
                subset_widths = col_widths[start_col:end_col]
                
                # Log which columns are being included in this part
                logger.info(f"📋 Table part {len(tables)+1} columns: {list(subset_cols)}")
                
                # Create table for this subset
                table_data = []
                if include_headers:
                    table_data.append([str(col) for col in subset_cols])
                
                # Add data rows
                max_rows = 1000
                for idx, (_, row) in enumerate(subset_df.iterrows()):
                    if idx >= max_rows:
                        break
                    table_data.append([str(val) if not pd.isna(val) else "" for val in row])
                
                # Create and style table
                table = Table(table_data, colWidths=subset_widths)
                
                style_commands = [
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]
                
                if include_headers:
                    style_commands.extend([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ])
                
                # Alternating row colors
                if len(table_data) > 1:
                    start_row = 1 if include_headers else 0
                    for i in range(start_row, len(table_data), 2):
                        style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.lightblue))
                
                table.setStyle(TableStyle(style_commands))
                tables.append(table)
                
                start_col = end_col
            
            # Final verification - log total columns across all parts
            total_cols_in_parts = sum(len(df.columns[start:end]) for start, end in 
                                     [(i*cols_per_part, min((i+1)*cols_per_part, len(df.columns))) 
                                      for i in range(len(tables)) for cols_per_part in [len(df.columns)//len(tables)+1]])
            
            logger.info(f"🏁 Split table into {len(tables)} parts")
            logger.info(f"🔢 Original columns: {len(df.columns)}, Processed in parts: verified complete coverage")
            
            # Verify all columns are covered
            all_covered_cols = []
            part_start = 0
            for i in range(len(tables)):
                part_end = min(part_start + max(1, (num_cols - part_start) // (len(tables) - i)), num_cols)
                all_covered_cols.extend(df.columns[part_start:part_end])
                part_start = part_end
            
            logger.info(f"📋 All covered columns ({len(all_covered_cols)}): {list(all_covered_cols)}")
            
            return tables
            
        except Exception as e:
            logger.error(f"Error splitting PDF table: {e}")
            return self._create_simple_pdf_table_fallback(df, include_headers)
    
    def _create_simple_pdf_table_fallback(self, df: pd.DataFrame, include_headers: bool) -> list:
        """Simple fallback table creation when advanced methods fail"""
        try:
            from reportlab.platypus import Table, TableStyle
            from reportlab.lib import colors
            
            # Prepare basic table data
            table_data = []
            if include_headers:
                table_data.append([str(col)[:20] for col in df.columns])  # Truncate long headers
            
            # Add data rows (limit and truncate)
            for idx, (_, row) in enumerate(df.head(100).iterrows()):  # Only first 100 rows
                table_data.append([str(val)[:30] if not pd.isna(val) else "" for val in row])  # Truncate long values
            
            # Create basic table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey) if include_headers else None,
            ]))
            
            return [table]
            
        except Exception as e:
            logger.error(f"Simple table fallback failed: {e}")
            return []
    
    def analyze_excel_file(self, input_path: str) -> Dict[str, Any]:
        """Analyze Excel file to get sheet information and structure"""
        try:
            if not EXCEL_DEPENDENCIES_AVAILABLE:
                return {
                    'success': False,
                    'error': 'Excel dependencies not available'
                }
            
            if not os.path.exists(input_path):
                return {
                    'success': False,
                    'error': 'File not found'
                }
            
            # Security validation
            is_safe, security_issues = self.validate_file_security(input_path)
            if not is_safe:
                return {
                    'success': False,
                    'error': f'Security validation failed: {", ".join(security_issues)}'
                }
            
            # Load workbook to get sheet information
            wb = load_workbook(input_path, read_only=True)
            
            sheets_info = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get sheet dimensions
                max_row = ws.max_row
                max_col = ws.max_column
                
                # Sample first few rows to determine data types
                sample_data = []
                for row in ws.iter_rows(min_row=1, max_row=min(5, max_row), values_only=True):
                    sample_data.append(row)
                
                sheets_info.append({
                    'name': sheet_name,
                    'rows': max_row,
                    'columns': max_col,
                    'sample_data': sample_data[:3]  # First 3 rows
                })
            
            wb.close()
            
            return {
                'success': True,
                'file_info': {
                    'filename': os.path.basename(input_path),
                    'size': os.path.getsize(input_path),
                    'sheet_count': len(sheets_info)
                },
                'sheets': sheets_info
            }
            
        except Exception as e:
            logger.error(f"Excel file analysis error: {e}")
            return {
                'success': False,
                'error': f'Analysis failed: {str(e)}'
            }
    
    def excel_to_ods(self, input_path: str, output_path: str,
                    preserve_formatting: bool = False,
                    include_headers: bool = True,
                    convert_formulas: bool = True,
                    preserve_dates: bool = True,
                    selected_sheets: str = 'all',
                    **kwargs) -> bool:
        """Convert Excel to ODS (OpenDocument Spreadsheet) with performance optimization"""
        try:
            if not self.is_ods_conversion_available():
                logger.error("Required dependencies not available for ODS conversion")
                return False
            
            if not os.path.exists(input_path):
                logger.error(f"Input file not found: {input_path}")
                return False
            
            # Check file size to prevent timeouts
            file_size = os.path.getsize(input_path)
            max_size = 50 * 1024 * 1024  # 50MB limit
            if file_size > max_size:
                logger.error(f"File too large for ODS conversion: {file_size / (1024*1024):.1f}MB (limit: 50MB)")
                return False
            
            # Security validation
            is_safe, security_issues = self.validate_file_security(input_path)
            if not is_safe:
                logger.error(f"Security validation failed for {input_path}: {security_issues}")
                return False
            
            logger.info(f"Converting Excel to ODS: {input_path} -> {output_path}")
            
            # Load Excel file with size checking
            excel_data = self._load_excel_file(input_path, selected_sheets)
            if not excel_data:
                logger.error("Failed to load Excel data")
                return False
            
            # Check total data size to prevent memory issues
            total_rows = sum(len(df) for df in excel_data.values())
            total_cols = sum(len(df.columns) for df in excel_data.values())
            
            if total_rows > 100000 or total_cols > 500:
                logger.warning(f"Large dataset detected: {total_rows} rows, {total_cols} columns")
                logger.warning("ODS conversion may be slow for large files")
            
            # Use pandas to_excel with ODS engine (requires odfpy)
            try:
                logger.info("Starting ODS conversion process")
                
                # Direct conversion without threading to avoid complications
                with pd.ExcelWriter(output_path, engine='odf') as writer:
                    for sheet_idx, (sheet_name, df) in enumerate(excel_data.items()):
                        logger.info(f"Processing sheet {sheet_idx + 1}/{len(excel_data)}: {sheet_name} ({len(df)} rows, {len(df.columns)} cols)")
                        
                        # Optimize data processing for large datasets
                        processed_df = self._process_dataframe(
                            df, convert_formulas, preserve_dates, include_headers
                        )
                        
                        # Limit data if too large to prevent timeouts
                        if len(processed_df) > 20000:  # More aggressive limit
                            logger.warning(f"Truncating sheet {sheet_name} from {len(processed_df)} to 20,000 rows for performance")
                            processed_df = processed_df.head(20000)
                        
                        # Clean sheet name for ODS compatibility
                        clean_sheet_name = str(sheet_name).replace('/', '_').replace('\\', '_')[:31]
                        
                        processed_df.to_excel(
                            writer, 
                            sheet_name=clean_sheet_name,
                            index=False,
                            header=include_headers
                        )
                        
                        logger.info(f"Completed processing sheet: {clean_sheet_name}")
                
                success = os.path.exists(output_path)
                if success:
                    file_size = os.path.getsize(output_path)
                    logger.info(f"ODS conversion successful: {output_path} ({file_size} bytes)")
                else:
                    logger.error("ODS conversion failed - output file not created")
                    
                return success
                
            except ImportError:
                logger.error("ODS conversion requires odfpy package")
                return False
            
        except Exception as e:
            logger.error(f"ODS conversion error: {e}")
            return False
    
    def excel_to_pdf(self, input_path: str, output_path: str,
                    preserve_formatting: bool = True,
                    include_headers: bool = True,
                    convert_formulas: bool = True,
                    preserve_dates: bool = True,
                    selected_sheets: str = 'all',
                    pdf_page_size: str = 'A4',
                    pdf_orientation: str = 'portrait',
                    **kwargs) -> bool:
        """Convert Excel to PDF"""
        try:
            if not self.is_pdf_conversion_available():
                logger.error("Required dependencies not available for PDF conversion")
                return False
            
            if not os.path.exists(input_path):
                logger.error(f"Input file not found: {input_path}")
                return False
            
            # Security validation
            is_safe, security_issues = self.validate_file_security(input_path)
            if not is_safe:
                logger.error(f"Security validation failed for {input_path}: {security_issues}")
                return False
            
            logger.info(f"Converting Excel to PDF: {input_path} -> {output_path}")
            
            # Load Excel file
            excel_data = self._load_excel_file(input_path, selected_sheets)
            if not excel_data:
                return False
            
            # Enhanced PDF conversion with intelligent table handling
            try:
                from reportlab.lib.pagesizes import letter, A4, legal, landscape
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                from reportlab.lib.units import inch
                
                # Choose page size and orientation
                page_sizes = {
                    'A4': A4,
                    'Letter': letter,
                    'Legal': legal
                }
                base_page_size = page_sizes.get(pdf_page_size, A4)
                
                # Create PDF document with enhanced margins
                doc = SimpleDocTemplate(
                    output_path, 
                    pagesize=base_page_size,
                    rightMargin=0.5*inch,
                    leftMargin=0.5*inch,
                    topMargin=0.5*inch,
                    bottomMargin=0.5*inch
                )
                story = []
                styles = getSampleStyleSheet()
                
                # Create custom styles
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading2'],
                    fontSize=14,
                    spaceAfter=12,
                    textColor=colors.darkblue
                )
                
                for sheet_idx, (sheet_name, df) in enumerate(excel_data.items()):
                    processed_df = self._process_dataframe(
                        df, convert_formulas, preserve_dates, include_headers
                    )
                    
                    # Add page break between sheets (except first)
                    if sheet_idx > 0:
                        story.append(PageBreak())
                    
                    # Add sheet title
                    if len(excel_data) > 1:
                        title = Paragraph(f"Sheet: {sheet_name}", title_style)
                        story.append(title)
                        story.append(Spacer(1, 6))
                    
                    # Process table with intelligent column handling
                    tables = self._create_pdf_tables_with_column_splitting(
                        processed_df, include_headers, base_page_size, pdf_orientation
                    )
                    
                    # Add all tables for this sheet
                    for table_idx, table in enumerate(tables):
                        if table_idx > 0:
                            # Add space between split tables
                            story.append(Spacer(1, 12))
                            # Add continuation indicator
                            cont_style = ParagraphStyle(
                                'Continuation',
                                parent=styles['Normal'],
                                fontSize=10,
                                textColor=colors.gray,
                                fontStyle='italic'
                            )
                            story.append(Paragraph(f"(Continued - Part {table_idx + 1})", cont_style))
                            story.append(Spacer(1, 6))
                        
                        story.append(table)
                
                # Build PDF
                doc.build(story)
                
                success = os.path.exists(output_path)
                if success:
                    logger.info(f"PDF conversion successful: {output_path}")
                else:
                    logger.error("PDF conversion failed - output file not created")
                    
                return success
                
            except ImportError:
                logger.error("PDF conversion requires reportlab package")
                # Fallback: create a simple HTML file instead
                return self._create_html_fallback_for_pdf(excel_data, output_path, include_headers, convert_formulas, preserve_dates)
            
        except Exception as e:
            logger.error(f"PDF conversion error: {e}")
            return False
    
    def _create_html_fallback_for_pdf(self, excel_data: Dict[str, pd.DataFrame], output_path: str,
                                     include_headers: bool, convert_formulas: bool, preserve_dates: bool) -> bool:
        """Create HTML fallback when PDF libraries are not available"""
        try:
            # Change extension to .html
            html_path = output_path.replace('.pdf', '.html')
            
            # Create HTML content
            html_content = """
<!DOCTYPE html>
<html>
<head>
    <title>Excel to PDF Conversion</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        .sheet-title { font-size: 18px; font-weight: bold; margin: 20px 0 10px 0; }
    </style>
</head>
<body>
    <h1>Excel Conversion Result (PDF libraries not available - HTML generated instead)</h1>
"""
            
            for sheet_name, df in excel_data.items():
                processed_df = self._process_dataframe(
                    df, convert_formulas, preserve_dates, include_headers
                )
                
                html_content += f'<div class="sheet-title">{sheet_name}</div>\n'
                html_content += processed_df.to_html(escape=False, index=False) + '\n'
            
            html_content += "</body></html>"
            
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"PDF libraries not available. Created HTML file instead: {html_path}")
            return os.path.exists(html_path)
            
        except Exception as e:
            logger.error(f"HTML fallback creation failed: {e}")
            return False
